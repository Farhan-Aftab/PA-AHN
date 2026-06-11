# ================================================================
# FILE: computational_cost_all_methods.py
# Clean computational-cost comparison for PA-AHN, SN, BN, and GN
# on CIFAR-10 / CIFAR-100.
#
# Purpose:
# - Measure clean training-time computational cost without analysis/logging overhead.
# - No per-batch BN/LN/IN weight logging.
# - No CSV logs.
# - No plots.
# - No checkpoint saving.
# - No file writing except one final Excel summary table.
# - No .cpu().numpy() weight logging.
#
# What is printed:
# - Method name and configuration
# - Parameter count
# - Epoch time per epoch
# - Average epoch time
# - Average batch time estimated from epoch time / number of batches
# - Peak GPU memory
#
# Important:
# - base_dir path is kept unchanged from thesis code.
# - This script assumes the following files already exist in the project:
#   models/resnet18_pa_ahn.py
#   models/resnet50_pa_ahn.py
#   norms/prior_anchored_adaptive_hybrid_normalization.py
#   models/resnet18_sn.py
#   models/resnet50_sn.py
#   norms/switchable_norm.py
#   models/resnet18_bn.py
#   models/resnet50_bn.py
#   models/resnet18_gn.py
#   models/resnet50_gn.py
# ================================================================

import os
import sys
import json
import random
import time
from dataclasses import dataclass, asdict
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
from torchvision import datasets, transforms
from dataclasses import dataclass, field
from typing import Optional

# ------------------------------------------------------------
# Keep thesis project path unchanged
# ------------------------------------------------------------
BASE_DIR = "/home/seecs/farhan/AHN_Thesis"
sys.path.append(BASE_DIR)

# PA-AHN imports
from models.resnet18_pa_ahn import ResNet18_AHN
from models.resnet50_pa_ahn import ResNet50_AHN
from norms.prior_anchored_adaptive_hybrid_normalization import AdaptiveHybridNorm2d

# SN imports
from models.resnet18_sn import ResNet18_SN
from models.resnet50_sn import ResNet50_SN
from norms.switchable_norm import SwitchableNorm2d

# BN imports
from models.resnet18_bn import ResNet18_BN
from models.resnet50_bn import ResNet50_BN

# GN imports
from models.resnet18_gn import ResNet18_GN
from models.resnet50_gn import ResNet50_GN


# =============================
# Configuration
# =============================
@dataclass
class Config:
    base_dir: str = "/home/seecs/farhan/AHN_Thesis"

    dataset_name: str = "cifar10"       # options: cifar10
    num_classes: int = 10               # options: cifar10

    # dataset_name: str = "cifar100"      # options: cifar100
    # num_classes: int = 100              # options: cifar100

    model_type: str = "resnet18"        # options: "resnet18"
    # model_type: str = "resnet50"        # options: "resnet50"

    # batch_size: int = 8                 # options: batch_size 8
    # batch_size: int = 32                # options: batch_size 32
    # batch_size: int = 64                # options: batch_size 64
    # batch_size: int = 128               # options: batch_size 128
    batch_size: int = 256               # options: batch_size 256

    # num_epochs: int = 100               # option: epochs 100
    # num_epochs: int = 50                # option: epochs 50
    num_epochs: int = 20                # option: epochs 20

    num_workers: int = 2
    momentum: float = 0.9
    weight_decay: float = 1e-4
    seed: int = 42

    # PA-AHN settings
    # temperature: float = 1.2
    # temperature: float = 1.0
    temperature: float = 0.8
    # temperature: float = 0.7
    # temperature: float = 0.3

    # Safe prior weights used to initialize PA-AHN base logits
    safe_bn_weight: float = 0.34
    safe_ln_weight: float = 0.33
    safe_in_weight: float = 0.33

    # L-Stab setting for final PA-AHN training cost.
    # This is part of the PA-AHN training method, not just logging.
    use_lstab: bool = True
    lstab_lambda_max: float = 0.00042
    anneal_lambda: bool = False
    use_temperature_scaling: bool = True

    # GN setting: 32 is standard for ResNet channels 64/128/256/512.
    gn_groups: int = 32

    # SN paper-style batch-average inference calibration.
    # For clean TRAINING computational-cost comparison, default is False.
    # Set True only if you intentionally want to include SN calibration overhead.
    use_sn_batch_average_inference_calibration: bool = True
    sn_batch_average_calibration_batches: Optional[int] = None

    # Measurement settings
    measure_gpu_memory: bool = True

    # Final output file. This is not a training log; it is the final clean summary table.
    # output_excel_name: str = "computational_cost_PA_AHN_SN_BN_GN{CFG.dataset_name.lower()}_{CFG.model_type}_batchsize_{CFG.batch_size}_{CFG.num_epochs}epochs.xlsx"

    output_excel_name: str = field(init=False)

    def __post_init__(self):
        self.output_excel_name = (
            f"computational_cost_PA_AHN_SN_BN_GN_"
            f"{self.dataset_name.upper()}_"
            f"{self.model_type.upper()}_"
            f"BATCHSIZE{self.batch_size}_"
            f"{self.num_epochs}EPOCHS.xlsx"
        )


CFG = Config()

# Same paper-style linear learning-rate scaling used in PA-AHN, SN, BN and GN code.
# effective batch 256 -> lr 0.1
# effective batch 32  -> lr 0.0125
# effective batch 8   -> lr 0.003125
CFG.lr = 0.1 * CFG.batch_size / 256


# =============================
# Utility functions
# =============================
def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def get_transforms(dataset_name: str):
    if dataset_name.lower() == "cifar10":
        mean = (0.4914, 0.4822, 0.4465)
        std = (0.2023, 0.1994, 0.2010)
    elif dataset_name.lower() == "cifar100":
        mean = (0.5071, 0.4867, 0.4408)
        std = (0.2675, 0.2565, 0.2761)
    else:
        raise ValueError(f"Unsupported dataset: {dataset_name}")

    transform_train = transforms.Compose([
        transforms.RandomHorizontalFlip(),
        transforms.RandomCrop(32, padding=4),
        transforms.ToTensor(),
        transforms.Normalize(mean, std),
    ])

    transform_test = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize(mean, std),
    ])
    return transform_train, transform_test


def get_dataloaders(cfg: Config):
    transform_train, transform_test = get_transforms(cfg.dataset_name)

    if cfg.dataset_name.lower() == "cifar10":
        train_dataset = datasets.CIFAR10(root="./data", train=True, download=True, transform=transform_train)
        test_dataset = datasets.CIFAR10(root="./data", train=False, download=True, transform=transform_test)
    elif cfg.dataset_name.lower() == "cifar100":
        train_dataset = datasets.CIFAR100(root="./data", train=True, download=True, transform=transform_train)
        test_dataset = datasets.CIFAR100(root="./data", train=False, download=True, transform=transform_test)
    else:
        raise ValueError(f"Unsupported dataset: {cfg.dataset_name}")

    train_loader = DataLoader(
        train_dataset,
        batch_size=cfg.batch_size,
        shuffle=True,
        num_workers=cfg.num_workers,
        pin_memory=True,
    )

    # Test loader is intentionally created for configuration consistency,
    # but the clean cost script does not evaluate test accuracy/loss.
    test_loader = DataLoader(
        test_dataset,
        batch_size=cfg.batch_size,
        shuffle=False,
        num_workers=cfg.num_workers,
        pin_memory=True,
    )
    return train_loader, test_loader


def count_parameters(model: nn.Module) -> int:
    # Fresh one-time count of trainable parameters.
    # This is not taken from previous logs or CSV files.
    return sum(p.numel() for p in model.parameters() if p.requires_grad)


def get_gpu_memory_mb() -> float:
    if torch.cuda.is_available():
        return float(torch.cuda.max_memory_allocated() / (1024 ** 2))
    return 0.0


def get_lambda_lstab(epoch: int, num_epochs: int, cfg: Config) -> float:
    if not cfg.use_lstab:
        return 0.0
    if cfg.anneal_lambda:
        return cfg.lstab_lambda_max * (1.0 - epoch / num_epochs)
    return cfg.lstab_lambda_max


def compute_ahn_entropy_loss(model: nn.Module) -> Tuple[torch.Tensor, int]:
    entropy_loss = None
    ahn_layers = 0

    for m in model.modules():
        if isinstance(m, AdaptiveHybridNorm2d):
            layer_entropy = m.compute_entropy()
            if entropy_loss is None:
                entropy_loss = layer_entropy
            else:
                entropy_loss = entropy_loss + layer_entropy
            ahn_layers += 1

    if entropy_loss is None:
        device = next(model.parameters()).device
        entropy_loss = torch.tensor(0.0, device=device)
    elif ahn_layers > 0:
        entropy_loss = entropy_loss / ahn_layers

    return entropy_loss, ahn_layers


def get_scaled_milestones(num_epochs: int) -> List[int]:
    ratios = [0.30, 0.60, 0.90]
    return sorted(set(max(1, int(round(num_epochs * r))) for r in ratios if int(round(num_epochs * r)) < num_epochs))


# ------------------------------------------------------------
# SN batch-average inference helpers kept for optional use.
# Default clean cost comparison does not call calibration.
# ------------------------------------------------------------
def start_batch_average_collection(model: nn.Module) -> None:
    for m in model.modules():
        if isinstance(m, SwitchableNorm2d):
            m.start_batch_average_collection()


def stop_batch_average_collection(model: nn.Module) -> None:
    for m in model.modules():
        if isinstance(m, SwitchableNorm2d):
            m.stop_batch_average_collection()


def finalize_batch_average(model: nn.Module) -> None:
    for m in model.modules():
        if isinstance(m, SwitchableNorm2d):
            m.finalize_batch_average()


def calibrate_sn_batch_average(model: nn.Module, train_loader: DataLoader, device: torch.device, max_batches: Optional[int] = None) -> None:
    was_training = model.training
    model.train()
    start_batch_average_collection(model)

    with torch.no_grad():
        for batch_idx, (inputs, _) in enumerate(train_loader):
            if max_batches is not None and batch_idx >= max_batches:
                break
            inputs = inputs.to(device)
            _ = model(inputs)

    stop_batch_average_collection(model)
    finalize_batch_average(model)

    if not was_training:
        model.eval()


# =============================
# Model builders
# =============================
def build_model(method_name: str, cfg: Config, device: torch.device) -> nn.Module:
    method_name = method_name.upper()
    temperature = cfg.temperature if cfg.use_temperature_scaling else 1.0

    if method_name == "PA-AHN":
        if cfg.model_type == "resnet18":
            model = ResNet18_AHN(num_classes=cfg.num_classes, temperature=temperature).to(device)
        elif cfg.model_type == "resnet50":
            model = ResNet50_AHN(num_classes=cfg.num_classes, temperature=temperature).to(device)
        else:
            raise ValueError("Invalid model_type")

        # Initialize the same safe prior in every AHN layer.
        safe_weights = torch.tensor(
            [cfg.safe_bn_weight, cfg.safe_ln_weight, cfg.safe_in_weight],
            dtype=torch.float32,
            device=device,
        )
        safe_weights = safe_weights / safe_weights.sum()

        for m in model.modules():
            if isinstance(m, AdaptiveHybridNorm2d):
                m.set_prior_weights(safe_weights)
        return model

    if method_name == "SN":
        if cfg.model_type == "resnet18":
            return ResNet18_SN(num_classes=cfg.num_classes).to(device)
        if cfg.model_type == "resnet50":
            return ResNet50_SN(num_classes=cfg.num_classes).to(device)
        raise ValueError("Invalid model_type")

    if method_name == "BN":
        if cfg.model_type == "resnet18":
            return ResNet18_BN(num_classes=cfg.num_classes).to(device)
        if cfg.model_type == "resnet50":
            return ResNet50_BN(num_classes=cfg.num_classes).to(device)
        raise ValueError("Invalid model_type")

    if method_name == "GN":
        if cfg.model_type == "resnet18":
            return ResNet18_GN(num_classes=cfg.num_classes, num_groups=cfg.gn_groups).to(device)
        if cfg.model_type == "resnet50":
            return ResNet50_GN(num_classes=cfg.num_classes, num_groups=cfg.gn_groups).to(device)
        raise ValueError("Invalid model_type")

    raise ValueError(f"Unsupported method: {method_name}")


# =============================
# Clean computational-cost runner
# =============================
def run_one_method(method_name: str, cfg: Config, train_loader: DataLoader, device: torch.device) -> Dict[str, object]:
    print("\n" + "=" * 78)
    print(f"STARTING METHOD: {method_name}")
    print("=" * 78)

    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.reset_peak_memory_stats(device)
        torch.cuda.synchronize()

    # Re-seed before each method for controlled initialization/data order.
    set_seed(cfg.seed)

    model = build_model(method_name, cfg, device)
    total_params = count_parameters(model)

    criterion = nn.CrossEntropyLoss()
    optimizer = optim.SGD(
        model.parameters(),
        lr=cfg.lr,
        momentum=cfg.momentum,
        weight_decay=cfg.weight_decay,
    )

    milestones = get_scaled_milestones(cfg.num_epochs)
    scheduler = optim.lr_scheduler.MultiStepLR(optimizer, milestones=milestones, gamma=0.1)

    method_config = {
        "Method": method_name,
        "Dataset": cfg.dataset_name,
        "Num_Classes": cfg.num_classes,
        "Model_Type": cfg.model_type,
        "Batch_Size": cfg.batch_size,
        "Num_Epochs": cfg.num_epochs,
        "Learning_Rate": cfg.lr,
        "Momentum": cfg.momentum,
        "Weight_Decay": cfg.weight_decay,
        "Scheduler_Milestones": milestones,
        "Parameter_Count": total_params,
    }

    if method_name == "PA-AHN":
        method_config.update({
            "Temperature": cfg.temperature if cfg.use_temperature_scaling else 1.0,
            "Use_LStab": cfg.use_lstab,
            "LStab_Lambda": cfg.lstab_lambda_max if cfg.use_lstab else 0.0,
            "Safe_BN_Weight": cfg.safe_bn_weight,
            "Safe_LN_Weight": cfg.safe_ln_weight,
            "Safe_IN_Weight": cfg.safe_in_weight,
        })
    elif method_name == "SN":
        method_config.update({
            "Use_SN_Batch_Average_Inference_Calibration": cfg.use_sn_batch_average_inference_calibration,
            "SN_Calibration_Batches": cfg.sn_batch_average_calibration_batches,
        })
    elif method_name == "GN":
        method_config.update({"GN_Groups": cfg.gn_groups})

    print("Configuration:")
    print(json.dumps(method_config, indent=2))
    print(f"{method_name} Parameter count: {total_params}")

    epoch_times: List[float] = []
    avg_batch_times: List[float] = []

    # Clean training loop:
    # - no train/test accuracy calculation
    # - no per-batch weight collection
    # - no .cpu().numpy() logging
    # - no CSV/plot/checkpoint/file logging
    for epoch in range(1, cfg.num_epochs + 1):
        model.train()

        if torch.cuda.is_available():
            torch.cuda.synchronize()
        epoch_start = time.perf_counter()

        lambda_lstab = get_lambda_lstab(epoch, cfg.num_epochs, cfg) if method_name == "PA-AHN" else 0.0

        for inputs, labels in train_loader:
            inputs = inputs.to(device, non_blocking=True)
            labels = labels.to(device, non_blocking=True)

            optimizer.zero_grad(set_to_none=True)
            outputs = model(inputs)
            task_loss = criterion(outputs, labels)

            if method_name == "PA-AHN" and cfg.use_lstab:
                entropy_loss, _ = compute_ahn_entropy_loss(model)
                loss = task_loss + lambda_lstab * entropy_loss
            else:
                loss = task_loss

            loss.backward()
            optimizer.step()

        scheduler.step()

        if method_name == "SN" and cfg.use_sn_batch_average_inference_calibration:
            calibrate_sn_batch_average(
                model=model,
                train_loader=train_loader,
                device=device,
                max_batches=cfg.sn_batch_average_calibration_batches,
            )

        if torch.cuda.is_available():
            torch.cuda.synchronize()
        epoch_time = time.perf_counter() - epoch_start

        avg_batch_time = epoch_time / max(1, len(train_loader))
        epoch_times.append(float(epoch_time))
        avg_batch_times.append(float(avg_batch_time))

        print(
            f"{method_name} | Epoch [{epoch}/{cfg.num_epochs}] | "
            f"EpochTime: {epoch_time:.2f}s | AvgBatchTime: {avg_batch_time:.6f}s"
        )

    peak_gpu_memory_mb = get_gpu_memory_mb() if cfg.measure_gpu_memory else 0.0
    average_epoch_time = float(np.mean(epoch_times)) if epoch_times else 0.0
    average_batch_time = float(np.mean(avg_batch_times)) if avg_batch_times else 0.0

    result = {
        "Method": method_name,
        "Dataset": cfg.dataset_name,
        "Model": cfg.model_type,
        "Batch_Size": cfg.batch_size,
        "Epochs": cfg.num_epochs,
        "Learning_Rate": cfg.lr,
        "Weight_Decay": cfg.weight_decay,
        "Momentum": cfg.momentum,
        "Scheduler_Milestones": str(milestones),
        "Parameter_Count": total_params,
        "Average_Epoch_Time_Sec": average_epoch_time,
        "Average_Batch_Time_Sec": average_batch_time,
        "Peak_GPU_Memory_MB": peak_gpu_memory_mb,
    }

    if method_name == "PA-AHN":
        result.update({
            "Temperature": cfg.temperature if cfg.use_temperature_scaling else 1.0,
            "Lambda": cfg.lstab_lambda_max if cfg.use_lstab else 0.0,
            "Extra_Setting": "PA-AHN clean train cost; no per-batch weight logs",
        })
    elif method_name == "SN":
        result.update({
            "Temperature": "N/A",
            "Lambda": "N/A",
            "Extra_Setting": f"SN clean train cost; calibration={cfg.use_sn_batch_average_inference_calibration}",
        })
    elif method_name == "BN":
        result.update({"Temperature": "N/A", "Lambda": "N/A", "Extra_Setting": "Standard BN baseline"})
    elif method_name == "GN":
        result.update({"Temperature": "N/A", "Lambda": "N/A", "Extra_Setting": f"GN groups={cfg.gn_groups}"})

    print("-" * 78)
    print(f"{method_name} CLEAN COMPUTATIONAL COST SUMMARY")
    print(f"Parameter count:       {result['Parameter_Count']}")
    print(f"Average epoch time:    {result['Average_Epoch_Time_Sec']:.2f}s")
    print(f"Average batch time:    {result['Average_Batch_Time_Sec']:.6f}s")
    print(f"Peak GPU memory:       {result['Peak_GPU_Memory_MB']:.2f} MB")
    print("-" * 78)

    # Free memory before next method.
    del model, optimizer, scheduler
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.synchronize()

    return result


def save_summary_excel(results: List[Dict[str, object]], cfg: Config) -> str:
    output_dir = os.path.join(cfg.base_dir, "results")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, cfg.output_excel_name)

    df = pd.DataFrame(results)

    # Professional column order for thesis/computation-cost table.
    columns = [
        "Method",
        "Dataset",
        "Model",
        "Batch_Size",
        "Epochs",
        "Learning_Rate",
        "Weight_Decay",
        "Momentum",
        "Scheduler_Milestones",
        "Temperature",
        "Lambda",
        "Parameter_Count",
        "Average_Epoch_Time_Sec",
        "Average_Batch_Time_Sec",
        "Peak_GPU_Memory_MB",
        "Extra_Setting",
    ]
    df = df[columns]

    # One final Excel file only. This is the clean computational-cost summary,
    # not the heavy training logs used in earlier thesis experiments.
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Clean Cost Summary", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Clean Cost Summary"]

        # Basic professional formatting for readability.
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = thin_border

        width_map = {
            "A": 14, "B": 12, "C": 12, "D": 10, "E": 8,
            "F": 14, "G": 12, "H": 10, "I": 22, "J": 12,
            "K": 10, "L": 16, "M": 20, "N": 20, "O": 18, "P": 38,
        }
        for col_letter, width in width_map.items():
            worksheet.column_dimensions[col_letter].width = width

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        # Add note sheet for interpretation.
        note_sheet = workbook.create_sheet("Interpretation Note")
        note_sheet["A1"] = "Clean Computational-Cost Measurement Note"
        note_sheet["A1"].font = Font(bold=True, size=14, color="1F4E78")
        note_sheet["A3"] = (
            "This file reports a cleaner computational-cost comparison because the script removes "
            "per-batch BN/LN/IN weight logging, CSV logging, plots, checkpoints, file writing, and "
            ".cpu().numpy() weight conversion. Average batch time is estimated from epoch time divided "
            "by the number of training batches, avoiding extra per-batch timing synchronization."
        )
        note_sheet["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        note_sheet.column_dimensions["A"].width = 110
        note_sheet.row_dimensions[3].height = 70

    return output_path


def main():
    set_seed(CFG.seed)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    print("\n" + "#" * 78)
    print("CLEAN COMPUTATIONAL-COST COMPARISON: PA-AHN, SN, BN, GN")
    print("#" * 78)
    print(f"Using device: {device}")
    print("Global configuration:")
    print(json.dumps(asdict(CFG), indent=2))
    print(f"Learning rate after linear scaling: {CFG.lr}")

    train_loader, _ = get_dataloaders(CFG)

    methods = ["PA-AHN", "SN", "BN", "GN"]
    results = []

    for method_name in methods:
        result = run_one_method(method_name, CFG, train_loader, device)
        results.append(result)

    output_excel = save_summary_excel(results, CFG)

    print("\n" + "#" * 78)
    print("FINAL CLEAN COMPUTATIONAL-COST TABLE")
    print("#" * 78)
    final_df = pd.DataFrame(results)
    display_cols = [
        "Method",
        "Parameter_Count",
        "Average_Epoch_Time_Sec",
        "Average_Batch_Time_Sec",
        "Peak_GPU_Memory_MB",
    ]
    print(final_df[display_cols].to_string(index=False))
    print(f"\nExcel summary saved to: {output_excel}")
    print("#" * 78)


if __name__ == "__main__":
    main()
